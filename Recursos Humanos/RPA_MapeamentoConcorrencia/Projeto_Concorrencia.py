"""
01 - Entrar no site dos concorrentes e capturar as informações das vagas abertas"

"""

##Dependências
import PySimpleGUI as sg 
from botcity.web import WebBot, Browser
from botcity.web import By
from botcity.web.browsers.chrome import default_options
from selenium.common.exceptions import NoSuchElementException


import openpyxl
import datetime
from botcity.web.util import element_as_select

##Variáveis de tempo

mes = str(datetime.datetime.now().strftime('%m'))
mes = int(mes)

semestre = ""

###Verificar o semestre da consulta

if mes == 1 or mes == 2 or mes ==  3 or mes ==  4:
    semestre = "1ºSemestre"

elif mes == 5 or mes == 6 or mes == 7:
    semestre = "2ºSemestre"

elif mes == 8 or mes == 9 or mes == 10 or mes == 11 or mes == 12:
    semestre = "1ºSemestre do próximo ano"

### Verificar o mês da consulta
if mes == 1:
    mes = "Janeiro"
elif mes == 2:
    mes = "Fevereiro"
elif mes == 3:
    mes = "Março"
elif mes == 4:
    mes = "Abril"
elif mes == 5:
    mes = "Maio"
elif mes == 6:
    mes = "Junho"
elif mes == 7:
    mes = "Julho"
elif mes == 8:
    mes = "Agosto"
elif mes == 9:
    mes = "Setembro"
elif mes == 10:
    mes = "Outubro"
elif mes == 11:
    mes = "Novembro"
elif mes == 12:
    mes = "Dezembro"

#Outras Variáveis

bot = WebBot()
hoje = str(datetime.datetime.now().strftime('%d.%m.%Y %Hh%Mm'))
bot.headless = False
bot.browser = Browser.CHROME
bot.driver_path = "chromedriver.exe"


## Listas

##Walljobs

lista_final_programas_walljobs = []
lista_final_clientes_walljobs = []
lista_final_links_walljobs = []

tipos_oportunidades_walljobs = []
anos_programas_walljobs = []

## Companhia de Estágios

lista_final_programas_ciaestagios = []
lista_final_clientes_ciaestagios = []
lista_final_links_ciaestagios = []

tipos_oportunidades_ciaestagios = []
anos_programas_ciaestagios = []

## EURECA

tipos_oportunidades_eureca = []
anos_programas_eureca = []

lista_final_programas_eureca = []
lista_final_links_eureca = []

##99jobs

lista_programas_estagios_99 = []
nomes_clientes_estagio_99 = []

lista_final_clientes_99 = []
lista_final_programas_99 = []

lista_final_links_99 = []

tipos_oportunidades_99 = []
anos_programas_99 = []

## ACROSS

nomes_programas_trainee_across = []
nomes_clientes_trainee_across = []
links_programas_across = []

nomes_programas_estagio_across = []
nomes_clientes_estagio_across = []

lista_final_programas_across = []
lista_final_clientes_across = []
lista_final_links_across = []

anos_programas_across = []
tipos_oportunidades_across = []

## MatchBox -> RETORNAR DEPOIS

lista_final_programas_matchbox = []
tipos_oportunidades_matchbox = []
anos_programas_matchbox = []
lista_final_links_matchbox = []

##Share RH

lista_final_programas_share = []
lista_final_clientes_share = []
lista_final_links_share = []

tipos_oportunidades_share = []
anos_programas_share = []


def tela_inicial():


    if __name__ == "__main__":


        sg.change_look_and_feel('Gray Gray Gray')

        tamanho_botao = (15,2)
    
        layout = [
            [sg.Column([[sg.Image(r'logo\logo-assinatura.png')]], justification='center')],
            [sg.Column([[sg.Text('Bem vindo a Automação de Concorrentes.',font=('Helvetica', 12 ,'bold'))]], justification='center')],
            [sg.Column([[sg.Text('Antes de começar, insira o seu login e senha do Cia de Estágios e Across',font=('Helvetica', 10, 'bold'))]], justification='center')],
            [sg.Column([[sg.Text('Atenção! Use cadastros e logins de uso pessoal e não corporativo para acessar os sites!',font=('Helvetica', 10, ))]], justification='center')],
            [sg.Column([[sg.Text('Essa medida é afim de evitar qualquer rastreamento por parte dos concorrentes.',font=('Helvetica', 10, ))]], justification='center')],
            [sg.Text('Login e senha Companhia de Estágios')],

            [sg.Text('Email de Login: '), sg.InputText(key="login_cia_estagios")],
            [sg.Text('Senha:             '), sg.InputText(password_char='*',key="senha_cia_estagios")],
            [sg.Text('')],    
             
            [sg.Text('Login e senha Across')],          
            [sg.Text('Email de Login: '), sg.InputText(key="login_across")],
            [sg.Text('Senha:             '), sg.InputText(password_char='*',key="senha_across")],

            [sg.Column([[sg.Text('Se você já escolheu a empresa, clique em COMEÇAR! para prosseguir')]], justification='center')],
            [sg.Text('')],
            [
            sg.Column([[sg.Button('Começar!', size=tamanho_botao, font=('Helvetica', 10, 'bold'))]], justification='center', element_justification='center'),

            ]]
        

        window = sg.Window('Automação Concorrência',layout, size=(700, 550))
        
        while True: 
            event, values = window.read()
            if event == sg.WIN_CLOSED:
                break   

            elif event == 'Começar!':

                window.close()
                email_cia_estagios = values['login_cia_estagios']
                senha_cia_estagios = values['senha_cia_estagios']

                email_across = values['login_across']
                senha_across = values['senha_across']
                
                across(email_across,senha_across)
                walljobs()
                ciaestagios(email_cia_estagios,senha_cia_estagios)
                matchbox()
                nineninejobs()
                eureca()                
                share_rh()
                joga_no_excel()


def walljobs ():

    ## Parte Trainee

    bot.browse('https://www.walljobs.com.br/vagas')

    bot.wait(5000)

    cookies_walljobs = bot.find_element("//a[@aria-label='dismiss cookie message']",By.XPATH)
    cookies_walljobs.click()

    
    tipo_vaga = bot.find_element('//*[@id="job_type"]',By.XPATH)
    tipo_vaga.click()

    bot.wait(6500)


    lista_tipos_trainee = bot.find_element("//ul[@role='listbox']/li[@data-value=4]",By.XPATH)
    lista_tipos_trainee.click()
    
    bot.wait(6500)

    qntd_vagas_walljobs_selector = bot.find_element("//div[contains(@class, 'filters-info')]/div/p",By.XPATH)
    qntd_vagas_walljobs = qntd_vagas_walljobs_selector.text
    qntd_vagas_walljobs = qntd_vagas_walljobs.split("de", 1)[1].strip()
    qntd_vagas_walljobs = qntd_vagas_walljobs.replace("south_east","")
    qntd_vagas_walljobs = int(qntd_vagas_walljobs)
    
    ##Resultados máximos de uma página = 30 por página

    if qntd_vagas_walljobs < 30:

        scrolladas = 2
    else:
         scrolladas = qntd_vagas_walljobs / 30
         scrolladas = int(scrolladas)
         print(scrolladas)


    for i in range(scrolladas):
            bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")
            bot.wait(2000)

    print(f'O número de scrolladas é: {scrolladas}')


    print(f'qntd de vagas = {qntd_vagas_walljobs}')

    bot.wait(3000)


    nome_programa_trainee_selector = bot.find_elements('//div[@class="display"]//h3',By.XPATH)
    nome_anunciante_trainee_selector = bot.find_elements('//div[@class="display"]//small',By.XPATH)
    url_selector_trainee_walljobs = bot.driver.find_elements_by_xpath('//a[@class="card-box-body"]')


    for i in range(len(nome_programa_trainee_selector)):
        
        nome_programa_trainee_walljobs = nome_programa_trainee_selector[i]
        nome_programa_trainee_walljobs = nome_programa_trainee_walljobs.text
        lista_final_programas_walljobs.append(nome_programa_trainee_walljobs)

        nome_cliente_trainee = nome_anunciante_trainee_selector[i * 4]
        nome_cliente_trainee = nome_cliente_trainee.text
        lista_final_clientes_walljobs.append(nome_cliente_trainee)
        
        url_walljobs_trainee = url_selector_trainee_walljobs[i]
        url_walljobs_trainee = url_walljobs_trainee.get_attribute('href')
        lista_final_links_walljobs.append(url_walljobs_trainee)

    print(len(lista_final_programas_walljobs))
    print(len(lista_final_clientes_walljobs))
    print(len(lista_final_links_walljobs))

    print(lista_final_links_walljobs)


    ###### Parte 1 Estágio

    bot.browse('https://www.walljobs.com.br/vagas')
    bot.wait(5000)

    tipo_vaga = bot.find_element('//*[@id="job_type"]',By.XPATH)
    tipo_vaga.click()

    bot.wait(5000)

    lista_tipos_estagio1 = bot.find_element("//ul[@role='listbox']/li[@data-value=3]",By.XPATH)
    lista_tipos_estagio1.click()

    bot.wait(5000)

    qntd_vagas_walljobs_selector = bot.find_element("//div[contains(@class, 'filters-info')]/div/p",By.XPATH)
    qntd_vagas_walljobs = qntd_vagas_walljobs_selector.text
    qntd_vagas_walljobs = qntd_vagas_walljobs.split("de", 1)[1].strip()
    qntd_vagas_walljobs = qntd_vagas_walljobs.replace("south_east","")
    qntd_vagas_walljobs = int(qntd_vagas_walljobs)

    if qntd_vagas_walljobs < 30:

        scrolladas = 2
    else:
         scrolladas = qntd_vagas_walljobs / 30
         scrolladas = int(scrolladas)
         print(scrolladas)


    for i in range(scrolladas):
            bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")
            bot.wait(2000)

    
    nome_programa_estagio_selector = bot.find_elements('//div[@class="display"]//h3',By.XPATH)
    nome_cliente_estagio_selector = bot.find_elements('//div[@class="display"]//small',By.XPATH)
    url_selector_estagio_walljobs = bot.driver.find_elements_by_xpath('//a[@class="card-box-body"]')


    for i in range(len(nome_programa_estagio_selector)):
        
        nome_programa_estagio_walljobs = nome_programa_estagio_selector[i]
        nome_programa_estagio_walljobs = nome_programa_estagio_walljobs.text
        lista_final_programas_walljobs.append(nome_programa_estagio_walljobs)

        nome_cliente_estagio = nome_cliente_estagio_selector[i * 4]
        nome_cliente_estagio = nome_cliente_estagio.text
        lista_final_clientes_walljobs.append(nome_cliente_estagio)

        url_walljobs_estagio = url_selector_estagio_walljobs[i]
        url_walljobs_estagio = url_walljobs_estagio.get_attribute('href')
        lista_final_links_walljobs.append(url_walljobs_estagio)

    # print(len(nomes_programas_estagio_walljobs))
    # print(len(nomes_clientes_estagio_walljobs))
    # print(len(urls_walljobs_estagio))

    ##Parte Estágio 2 

    bot.browse('https://www.walljobs.com.br/vagas')

    tipo_vaga = bot.find_element('//*[@id="job_type"]',By.XPATH)
    tipo_vaga.click()
        
    bot.wait(5000)
    
    lista_tipos_estagio2 = bot.find_element("//ul[@role='listbox']/li[@data-value=2]",By.XPATH)
    lista_tipos_estagio2.click()
    
    bot.wait(5000)

    qntd_vagas_walljobs_selector = bot.find_element("//div[contains(@class, 'filters-info')]/div/p",By.XPATH)
    qntd_vagas_walljobs = qntd_vagas_walljobs_selector.text
    qntd_vagas_walljobs = qntd_vagas_walljobs.split("de", 1)[1].strip()
    qntd_vagas_walljobs = qntd_vagas_walljobs.replace("south_east","")
    qntd_vagas_walljobs = int(qntd_vagas_walljobs)

    if qntd_vagas_walljobs < 30:

        scrolladas = 2
    else:
         scrolladas = qntd_vagas_walljobs / 30
         scrolladas = int(scrolladas)
         print(scrolladas)


    for i in range(scrolladas):
            bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")
            bot.wait(2000)
    

    nome_programa_estagio_selector = bot.find_elements('//div[@class="display"]//h3',By.XPATH)
    nome_cliente_estagio_selector = bot.find_elements('//div[@class="display"]//small',By.XPATH)
    url_selector_estagio_walljobs = bot.driver.find_elements_by_xpath('//a[@class="card-box-body"]')


    for i in range(len(nome_programa_estagio_selector)):
        
        nome_programa_estagio_walljobs = nome_programa_estagio_selector[i]
        nome_programa_estagio_walljobs = nome_programa_estagio_walljobs.text
        lista_final_programas_walljobs.append(nome_programa_estagio_walljobs)

        nome_cliente_estagio = nome_cliente_estagio_selector[i * 4]
        nome_cliente_estagio = nome_cliente_estagio.text
        lista_final_clientes_walljobs.append(nome_cliente_estagio)

        url_walljobs_estagio = url_selector_estagio_walljobs[i]
        url_walljobs_estagio = url_walljobs_estagio.get_attribute('href')
        lista_final_links_walljobs.append(url_walljobs_estagio)

    
    for i in range(len(lista_final_programas_walljobs)):
        
        tipo_programa = lista_final_programas_walljobs[i]
    
        if "Trainee" in tipo_programa or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_walljobs.append(tipo_programa)

        elif "Estágio" in tipo_programa or "Internship" in tipo_programa or "estagio" in tipo_programa or "ESTAGIO" in tipo_programa or "estágio" in tipo_programa:
            tipo_programa = "Estágio"
            tipos_oportunidades_walljobs.append(tipo_programa)

        else:
             tipo_programa = "Outros"
             tipos_oportunidades_walljobs.append(tipo_programa)
        
        ano_programa = lista_final_programas_walljobs[i]

        if "2023" in ano_programa:
             ano_programa = "2023"
             anos_programas_walljobs.append(ano_programa)
        elif "2024" in ano_programa:
             ano_programa = "2024"
             anos_programas_walljobs.append(ano_programa)

        elif "2025" in ano_programa:
             ano_programa = "2025"
             anos_programas_walljobs.append(ano_programa)        
        else:
             ano_programa = "Ano atual da pesquisa"
             anos_programas_walljobs.append(ano_programa)            

def ciaestagios(email_cia_estagios,senha_cia_estagios):
      
      bot.browse('https://app.ciadeestagios.com.br/')

      entre_aqui = bot.find_element("//button[@type='button' and span[text()='Entre aqui']]",By.XPATH)
      entre_aqui.click()
      bot.wait(4000)

      usuario_ciaestagios = bot.find_element("//*[@id='identity']",By.XPATH)
      usuario_ciaestagios.send_keys(email_cia_estagios)

      senha_ciaestagios = bot.find_element('//*[@id="password"]',By.XPATH)
      senha_ciaestagios.send_keys(senha_cia_estagios)

      entre_aqui = bot.find_element("//button[@type='button' and span[text()='Enviar']]",By.XPATH)
      entre_aqui.click()
      entre_aqui.click()

      bot.wait(6000)

      bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")
      bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")

      nome_programa_selector = bot.find_elements('//md-card[@class="_md"]//md-card-header-text/span[@class="md-title ng-binding"]',By.XPATH)
      nome_cliente_selector = bot.find_elements('//md-card[@class="_md"]//md-card-header-text/span[@class="md-subhead ng-binding"]',By.XPATH)

      link_programa_selector1 = bot.driver.find_elements_by_xpath('//md-card[starts-with(@id, "PE")]')
      link_programa_selector2 = bot.driver.find_elements_by_xpath('//md-card[starts-with(@id, "DC")]')
      link_programa_selector3 = bot.driver.find_elements_by_xpath('//md-card[starts-with(@id, "FC")]')


      for i in range(len(nome_programa_selector)):
      
        nome_programa_ciaestagios = nome_programa_selector[i].text
        lista_final_programas_ciaestagios.append(nome_programa_ciaestagios)

        nome_cliente_ciaestagios = nome_cliente_selector[i].text
        lista_final_clientes_ciaestagios.append(nome_cliente_ciaestagios)

    
      for i in range(len(link_programa_selector1)):

        link_programa = link_programa_selector1[i]
        link_programa = link_programa.get_attribute('id') 
        link_programa = link_programa[2:]

        vaga_ou_empresa = len(link_programa)

        if vaga_ou_empresa > 4:

            link_programa_final_ciadeestagios = f'https://app.ciadeestagios.com.br/vagas/{link_programa}'
            lista_final_links_ciaestagios.append(link_programa_final_ciadeestagios)

        if vaga_ou_empresa <= 4:

            link_programa_final_ciadeestagios = f'https://app.ciadeestagios.com.br/empresas/{link_programa}'
            lista_final_links_ciaestagios.append(link_programa_final_ciadeestagios)
        

      for i in range(len(link_programa_selector2)):
            
        link_programa = link_programa_selector2[i]
        link_programa = link_programa.get_attribute('id') 
        link_programa = link_programa[2:]

        vaga_ou_empresa = len(link_programa)

        if vaga_ou_empresa > 4:

            link_programa_final_ciadeestagios = f'https://app.ciadeestagios.com.br/vagas/{link_programa}'
            lista_final_links_ciaestagios.append(link_programa_final_ciadeestagios)

        if vaga_ou_empresa <= 4:

            link_programa_final_ciadeestagios = f'https://app.ciadeestagios.com.br/empresas/{link_programa}'
            lista_final_links_ciaestagios.append(link_programa_final_ciadeestagios)

      for i in range(len(link_programa_selector3)):
            
        link_programa = link_programa_selector3[i]
        link_programa = link_programa.get_attribute('id') 
        link_programa = link_programa[2:]

        vaga_ou_empresa = len(link_programa)

        if vaga_ou_empresa > 4:

            link_programa_final_ciadeestagios = f'https://app.ciadeestagios.com.br/vagas/{link_programa}'
            lista_final_links_ciaestagios.append(link_programa_final_ciadeestagios)

        if vaga_ou_empresa <= 4:

            link_programa_final_ciadeestagios = f'https://app.ciadeestagios.com.br/empresas/{link_programa}'
            lista_final_links_ciaestagios.append(link_programa_final_ciadeestagios)


      for i in range(len(lista_final_programas_ciaestagios)):

        tipo_programa = lista_final_programas_ciaestagios[i]
    
        if "Trainee" in tipo_programa or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_ciaestagios.append(tipo_programa)

        elif "Estágio" in tipo_programa or "Internship" in tipo_programa or "estagio" in tipo_programa or "ESTAGIO" in tipo_programa or "estágio" in tipo_programa:
            tipo_programa = "Estágio"
            tipos_oportunidades_ciaestagios.append(tipo_programa)

        else:
             tipo_programa = "Outros"
             tipos_oportunidades_ciaestagios.append(tipo_programa)
        
        ano_programa = lista_final_programas_ciaestagios[i]

        if "2023" in ano_programa:
             ano_programa = "2023"
             anos_programas_ciaestagios.append(ano_programa)
        elif "2024" in ano_programa:
             ano_programa = "2024"
             anos_programas_ciaestagios.append(ano_programa)    
        elif "2025" in ano_programa:
             ano_programa = "2025"
             anos_programas_ciaestagios.append(ano_programa)    
        else:
             ano_programa = "Ano atual da pesquisa"
             anos_programas_ciaestagios.append(ano_programa)

def eureca():

    bot.browse('https://oportunidades.eureca.me/')

    nome_programa_eureca_selector = bot.find_elements("//div[starts-with(@class, 'card-body')]/h3",By.XPATH)
    nome_cliente_eureca_selector = bot.find_elements("//div[@class='MuiCardContent-root']/p",By.XPATH)
    link_programa_eureca_selector = bot.driver.find_elements_by_xpath("//a[@class='h-64 w-full max-w-md']")


    for i in range(len(nome_programa_eureca_selector)):
         
        nome_programa_eureca = nome_programa_eureca_selector[i]
        nome_programa_eureca = nome_programa_eureca.text
        lista_final_programas_eureca.append(nome_programa_eureca)

        # CANCELADO POR BLOQUEAREM NO SELENIUM A INFORMAÇÃO

        # nome_cliente_eureca = nome_cliente_eureca_selector[i]
        # nome_cliente_eureca = nome_cliente_eureca.text
        # nomes_clientes_eureca.append(nome_cliente_eureca)

        link_programa_eureca = link_programa_eureca_selector[i]
        link_programa_eureca = link_programa_eureca.get_attribute('href')
        
        print(link_programa_eureca)
        lista_final_links_eureca.append(link_programa_eureca)


    for i in range(len(lista_final_programas_eureca)):
         
        tipo_programa = lista_final_programas_eureca[i]
    
        if "Trainee" or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_eureca.append(tipo_programa)

        elif "Estágio" or "Internship" or "estagio" or "ESTAGIO" or "estágio " in tipo_programa:
            tipo_programa = "Estágio"
            tipos_oportunidades_eureca.append(tipo_programa)

        else:
             tipo_programa = "Outros"
             tipos_oportunidades_eureca.append(tipo_programa)
        
        ano_programa = lista_final_programas_eureca[i]

        if "2023" in ano_programa:
             ano_programa = "2023"
             anos_programas_eureca.append(ano_programa)
        elif "2024" in ano_programa:
             ano_programa = "2024"
             anos_programas_eureca.append(ano_programa)       
        elif "2025" in ano_programa:
             ano_programa = "2025"
             anos_programas_eureca.append(ano_programa)        
        else:
             ano_programa = "Ano atual da pesquisa"
             anos_programas_eureca.append(ano_programa)

def across(email_across,senha_across):
     
     bot.browse("https://login.across.jobs/?pUrlRedirecionadaPortal=/")

     usuario = bot.find_element('//*[@id="NomeUsuario"]',By.XPATH)
     usuario.send_keys(email_across)
     bot.wait(2000)

     senha = bot.find_element('//*[@id="Senha"]',By.XPATH)
     senha.send_keys(senha_across)
     bot.wait(2000)

     clique_login = bot.find_element('//*[@id="btnLoginPrimario"]',By.XPATH)
     clique_login.click()
     bot.wait(2000)
     

     oportunidades = bot.find_element("//*[@id='liMenuProgramaTeste']/a/span",By.XPATH)
     oportunidades.click()
     bot.wait(3500)

     
     bot.browse("https://portal.across.jobs/programa/MostraTipo?pTipo=2")

     bot.wait(3500)

    #  bot.driver.find_element_by_css_selector("#formListagemProgramas > div > div.col-sm-12 > div > label:nth-child(1)").click()
     bot.wait(10000)

     sites_across_trainee_selector = bot.find_elements("//a[contains(@class, 'btn') and contains(@class, 'btn-primary') and contains(text(), 'Inscreva-se')]",By.XPATH)
     nome_programa_trainee_across_selector = bot.find_elements("//div[@class='caption']/p",By.XPATH)
     nome_cliente_trainee_across_selector = bot.find_elements("//span[@class='panel-title']/strong",By.XPATH)


     for i in range(len(sites_across_trainee_selector)):
        
        nome_programa_trainee_across = nome_programa_trainee_across_selector[i]
        nome_programa_trainee_across = nome_programa_trainee_across.text
        lista_final_programas_across.append(nome_programa_trainee_across)
 
        nome_cliente_trainee_across = nome_cliente_trainee_across_selector[i]
        nome_cliente_trainee_across = nome_cliente_trainee_across.text
        lista_final_clientes_across.append(nome_cliente_trainee_across)

     bot.wait(3000)

     for i in range(len(sites_across_trainee_selector)):

        bot.wait(4000)

        sites_across_trainee = sites_across_trainee_selector[i]
        sites_across_trainee.click()

        bot.wait(2000)

        try:

            bot.wait(2000)

            captura_link_site_selector = bot.driver.find_elements_by_xpath("//p[@class='acesseosite1']/a")
            
            captura_link_site = captura_link_site_selector[0]
            
            captura_link_site = captura_link_site.get_attribute('href')

            links_programas_across.append(captura_link_site)
            bot.back()


        except:
            bot.back()
            links_programas_across.append("Não tem link")


     ## Etapa Estágio

     bot.browse("https://portal.across.jobs/programa/MostraTipo?pTipo=1")

     bot.wait(5000)

    #  bot.driver.find_element_by_css_selector("#formListagemProgramas > div > div.col-sm-12 > div > label:nth-child(1)").click()
     bot.wait(6000)

     sites_across_estagio_selector = bot.find_elements("//a[contains(@class, 'btn') and contains(@class, 'btn-primary') and contains(text(), 'Inscreva-se')]",By.XPATH)
     nome_programa_estagio_across_selector = bot.find_elements("//div[@class='caption']/p",By.XPATH)
     nome_cliente_estagio_across_selector = bot.find_elements("//span[@class='panel-title']/strong",By.XPATH)


     for i in range(len(sites_across_estagio_selector)):
        
        nome_programa_estagio_across = nome_programa_estagio_across_selector[i]
        nome_programa_estagio_across = nome_programa_estagio_across.text
        lista_final_programas_across.append(nome_programa_estagio_across)


        nome_cliente_estagio_across = nome_cliente_estagio_across_selector[i]
        nome_cliente_estagio_across = nome_cliente_estagio_across.text
        lista_final_clientes_across.append(nome_cliente_estagio_across)

     bot.wait(2000)
    
     for i in range(len(sites_across_estagio_selector)):

        sites_across_estagio_selector2 = bot.find_elements("//a[contains(@class, 'btn') and contains(@class, 'btn-primary') and contains(text(), 'Inscreva-se')]",By.XPATH)
        sites_across_estagio = sites_across_estagio_selector2[i]
        sites_across_estagio.click()

        try:

            bot.wait(2000)

            captura_link_site_selector = bot.driver.find_elements_by_xpath("//p[@class='acesseosite1']/a")
            
            captura_link_site = captura_link_site_selector[0]
            
            captura_link_site = captura_link_site.get_attribute('href')

            links_programas_across.append(captura_link_site)

            bot.back()
            bot.wait(3000)


        except:
            links_programas_across.append("Não tem link")
            bot.back()
            bot.wait(3000)


     for i in range(len(lista_final_programas_across)):

        tipo_programa = lista_final_programas_across[i]

        if "Trainee" or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_across.append(tipo_programa)

        elif "Estágio" or "Internship" or "estagio" or "ESTAGIO" or "estágio " in tipo_programa:
            tipo_programa = "Estágio"
            tipos_oportunidades_across.append(tipo_programa)

        else:
                tipo_programa = "Outros"
                tipos_oportunidades_across.append(tipo_programa)
        
        ano_programa = lista_final_programas_across[i]

        if "2023" in ano_programa:
                ano_programa = "2023"
                anos_programas_across.append(ano_programa)
        elif "2024" in ano_programa:
                ano_programa = "2024"
                anos_programas_across.append(ano_programa)       
        elif "2025" in ano_programa:
                ano_programa = "2025"
                anos_programas_across.append(ano_programa)        
        else:
                ano_programa = "Ano atual da pesquisa"
                anos_programas_across.append(ano_programa)

    
     print(len(lista_final_programas_across))
     print(len(lista_final_clientes_across))
     print(len(links_programas_across))

     print(lista_final_programas_across)
     print(lista_final_clientes_across)
     print(links_programas_across)


# Texto já comentado

     print(nomes_programas_trainee_across)
     print(nomes_clientes_trainee_across)
     print(len(nomes_programas_trainee_across))
     print(len(nomes_clientes_trainee_across))



     print(nomes_programas_estagio_across)
     print(nomes_clientes_estagio_across)
     print(len(nomes_programas_estagio_across))
     print(len(nomes_clientes_estagio_across))
     print(len(links_programas_across))




     bot.wait(10000)

def matchbox():
     
    bot.browse("https://linktr.ee/carreirasmatch")

    botao_cookie = bot.find_element("//button[@id='onetrust-accept-btn-handler']",By.XPATH)
    botao_cookie.click()
    bot.wait(2000)
    bot.refresh()
    bot.wait(5000)                                                                        
    nome_programa_matchbox_selector = bot.find_elements("//*[@id='__next']/div[2]/div/div[1]/div[3]//p",By.XPATH)
    links_programas_matchbox_selector = bot.driver.find_elements_by_xpath('//*[@id="__next"]/div[2]/div/div[1]/div[3]//a')

    for i in range(len(nome_programa_matchbox_selector)):
     
        nome_programa_matchbox = nome_programa_matchbox_selector[i]
        nome_programa_matchbox = nome_programa_matchbox.text

        lista_final_programas_matchbox.append(nome_programa_matchbox)

        links_programas_matchbox = links_programas_matchbox_selector[i]
        links_programas_matchbox = links_programas_matchbox.get_attribute('href')

        lista_final_links_matchbox.append(links_programas_matchbox)

        for i in range(len(lista_final_programas_matchbox)):
         
         tipo_programa = lista_final_programas_matchbox[i]

    
         if "Trainee" or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_matchbox.append(tipo_programa)

         elif "Estágio" or "Internship" or "estagio" or "ESTAGIO" or "estágio " in tipo_programa:

            tipo_programa = "Estágio"
            tipos_oportunidades_matchbox.append(tipo_programa)

         else:
             tipo_programa = "Outros"
             tipos_oportunidades_matchbox.append(tipo_programa)
        
         ano_programa = lista_final_programas_matchbox[i]


         if "2023" in ano_programa:
             ano_programa = "2023"
             anos_programas_matchbox.append(ano_programa)
         elif "2024" in ano_programa:
             ano_programa = "2024"
             anos_programas_matchbox.append(ano_programa)  
         elif "2025" in ano_programa:
             ano_programa = "2025"
             anos_programas_matchbox.append(ano_programa)   
         else:
             ano_programa = "Ano atual da pesquisa"
             anos_programas_matchbox.append(ano_programa)

def nineninejobs():
     
    bot.browse("https://99jobs.com/opportunities/filtered_search?utf8=%E2%9C%93&search%5Bterm%5D=Estagio")

    bot.wait(3000)

    check_estagio_99 = bot.find_element('//label[contains(@for, "search_level_3") and contains(text(), "Estágio")]',By.XPATH)
    check_estagio_99.click()

    bot.wait(3000) 

    vermais_99jobs = bot.find_element("body > div.search-opportunities-page > div.search-opportunities-main > div > div.load-more-opportunitites > button",By.CSS_SELECTOR)
    

    while vermais_99jobs.is_enabled:
          try:
            vermais_99jobs.click()
          except:
               break
                                          
    num_vagas = bot.find_element("//span[@id='text-total-opportunities']",By.XPATH)
    num_vagas = num_vagas.text
    num_vagas = int(num_vagas)
    num_vagas_for = num_vagas * 2

    for i in range(2,num_vagas_for + 1,2):
         programa_estagio_selector99 = bot.find_element(f'body > div.search-opportunities-page > div.search-opportunities-main > div > div.opportunities-list > a:nth-child({str(i)}) > div > h1',By.CSS_SELECTOR)
         programa_estagio99 = programa_estagio_selector99.text
         lista_final_programas_99.append(programa_estagio99)
         
         programa_estagio_cliente_selector99 = bot.find_element(f"body > div.search-opportunities-page > div.search-opportunities-main > div > div.opportunities-list > a:nth-child({str(i)}) > div > div.opportunity-card-footer > div > div.opportunity-company-infos > h2",By.CSS_SELECTOR)
         programa_estagio_cliente_99 = programa_estagio_cliente_selector99.text
         lista_final_clientes_99.append(programa_estagio_cliente_99)


    for i in range(num_vagas):

         links_programas_99_selector = bot.driver.find_elements_by_xpath('//a[@class="opportunity-card"]')
         links_programas_99 = links_programas_99_selector[i]
         links_programas_99 = links_programas_99.get_attribute('href')
         lista_final_links_99.append(links_programas_99)

    # Tempo para esperar adicionar as coisas na lista e ir para trainee
    bot.wait(6000)

    bot.browse("https://99jobs.com/opportunities/filtered_search?utf8=%E2%9C%93&search%5Bterm%5D=trainee")

    bot.wait(3000) 
    check_estagio_99 = bot.find_element('//label[contains(@for, "search_level_4") and contains(text(), "Trainee")]',By.XPATH)
    check_estagio_99.click()
    bot.wait(3000)

#######################################

    num_vagas = bot.find_element("//span[@id='text-total-opportunities']",By.XPATH)
    num_vagas = num_vagas.text
    num_vagas = int(num_vagas)
    num_vagas_for = num_vagas * 2

    for i in range(2,num_vagas_for + 1,2):
         
         programa_trainee_selector_99 = bot.find_element(f'body > div.search-opportunities-page > div.search-opportunities-main > div > div.opportunities-list > a:nth-child({str(i)}) > div > h1',By.CSS_SELECTOR)
         programa_trainee99 = programa_trainee_selector_99.text
         lista_final_programas_99.append(programa_trainee99)
         
         programa_trainee_cliente_99_selector = bot.find_element(f"body > div.search-opportunities-page > div.search-opportunities-main > div > div.opportunities-list > a:nth-child({str(i)}) > div > div.opportunity-card-footer > div > div.opportunity-company-infos > h2",By.CSS_SELECTOR)
         programa_estagio_cliente_99 = programa_trainee_cliente_99_selector.text
         lista_final_clientes_99.append(programa_estagio_cliente_99)

    for i in range(num_vagas):

        links_programas_99_selector = bot.driver.find_elements_by_xpath('//a[@class="opportunity-card"]')
        links_programas_99 = links_programas_99_selector[i]
        links_programas_99 = links_programas_99.get_attribute('href')
        lista_final_links_99.append(links_programas_99)


    for i in range(len(lista_final_programas_99)):
        
        tipo_programa = lista_final_programas_99[i]
    
        if "Trainee" in tipo_programa or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_99.append(tipo_programa)

        elif "Estágio" in tipo_programa or "Internship" in tipo_programa or "estagio" in tipo_programa or "ESTAGIO" in tipo_programa or "estágio" in tipo_programa:
            tipo_programa = "Estágio"
            tipos_oportunidades_99.append(tipo_programa)

        else:
             tipo_programa = "Outros"
             tipos_oportunidades_99.append(tipo_programa)
        
        ano_programa = lista_final_programas_99[i]

        if "2023" in ano_programa:
             ano_programa = "2023"
             anos_programas_99.append(ano_programa)
        elif "2024" in ano_programa:
             ano_programa = "2024"
             anos_programas_99.append(ano_programa)       
        elif "2025" in ano_programa:
             ano_programa = "2025"
             anos_programas_99.append(ano_programa)        
        else:
             ano_programa = "Ano atual da pesquisa"
             anos_programas_99.append(ano_programa)      


    print(lista_final_programas_99)
    print(lista_final_clientes_99)
    print(lista_final_links_99)
    print()
    print()
    print(len(lista_final_programas_99))
    print(len(lista_final_clientes_99))
    print(len(lista_final_links_99))

def share_rh():

    bot.browse("https://sharepeoplehub.inhire.app/vagas")

    bot.wait(10000)

    bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")
    bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")

    bot.wait(5000)

    nomes_vagas_share_selector = bot.find_elements("//div[@data-component-name='PublicPageBody']//li",By.XPATH)
    
    for i in range(len(nomes_vagas_share_selector)):

        nomes_vagas_share = nomes_vagas_share_selector[i]
        nomes_programas_share = nomes_vagas_share.text
        lista_final_programas_share.append(nomes_programas_share)

        link_vagas_share_selector = bot.driver.find_elements_by_xpath("//div[@data-component-name='PublicPageBody']//li//a")
        link_vagas_share = link_vagas_share_selector[i]
        link_vagas_share = link_vagas_share.get_attribute('href')
        
        lista_final_links_share.append(link_vagas_share)


    for i in range(len(lista_final_links_share)):

     bot.browse(lista_final_links_share[i])
     bot.wait(5000)

     nomes_clientes_share_selector = bot.find_element('//div[@data-component-name="HtmlParser"]//p',By.XPATH)

     nomes_clientes_share = nomes_clientes_share_selector
     nomes_clientes_share = nomes_clientes_share.text
     nomes_clientes_share = nomes_clientes_share[37:]
     
     indice_virgula = nomes_clientes_share.find(",")
     nomes_clientes_share = nomes_clientes_share[0:indice_virgula]
     
     lista_final_clientes_share.append(nomes_clientes_share)
     print(nomes_clientes_share)

    for i in range(len(lista_final_programas_share)):
        
        tipo_programa = lista_final_programas_share[i]
    
        if "Trainee" in tipo_programa or "trainee" in tipo_programa:
            tipo_programa = "Trainee"
            tipos_oportunidades_share.append(tipo_programa)

        elif "Estágio" in tipo_programa or "Internship" in tipo_programa or "estagio" in tipo_programa or "ESTAGIO" in tipo_programa or "estágio" in tipo_programa:
            tipo_programa = "Estágio"
            tipos_oportunidades_share.append(tipo_programa)

        else:
             tipo_programa = "Outros"
             tipos_oportunidades_share.append(tipo_programa)
        
        ano_programa = lista_final_programas_share[i]

        if "2023" in ano_programa:
             ano_programa = "2023"
             anos_programas_share.append(ano_programa)
        elif "2024" in ano_programa:
             ano_programa = "2024"
             anos_programas_share.append(ano_programa)
        elif "2025" in ano_programa:
             ano_programa = "2025"
             anos_programas_share.append(ano_programa) 
        else:
             ano_programa = "Ano atual da pesquisa"
             anos_programas_share.append(ano_programa)      

def joga_no_excel():

    workbook = openpyxl.Workbook()

    sheet_mapeamento = workbook.create_sheet(title="Mapeamento")

    sheet_mapeamento.append({'A': 'Consultoria', 
                             'B': 'Cliente',
                             'C':'Tipo de Seleção',
                             'D':'Nível Complexidade',
                             'E':'Nome do programa',
                             'F':'Semestre','G':'Ano',
                             'H':'Diversidade',
                             'I':'Obs',
                             'J':'Link',
                             'K':'Mês de Consulta',
                             'L':'Fizemos Proposta?',
                             'M':'Nº Proposta',
                             'N':'Pq perdemos? (Moskit)'
                             })
                             

    
    ## Caso Walljobs
    for i in range(len(lista_final_clientes_walljobs)):
        
        sheet_mapeamento.append({
            'A': 'WallJobs',
            'B': lista_final_clientes_walljobs[i],
            'C': tipos_oportunidades_walljobs[i],
            'D': '',
            'E': lista_final_programas_walljobs[i],
            'F': semestre,
            'G': anos_programas_walljobs[i],
            'H': '',
            'I': '',
            'J': lista_final_links_walljobs[i],
            'K': mes,
            'L': '',
            'M': '',
            'N': ''
        })


        ##Caso COMPANHIA DE ESTÁGIOS

    for i in range(len(lista_final_programas_ciaestagios)):
        
        sheet_mapeamento.append({
            'A': 'Companhia de Estágios',
            'B': lista_final_clientes_ciaestagios[i],
            'C': tipos_oportunidades_ciaestagios[i],
            'D': '',
            'E': lista_final_programas_ciaestagios[i],
            'F': semestre,
            'G': anos_programas_ciaestagios[i],
            'H': '',
            'I': '',
            'J': lista_final_links_ciaestagios[i],
            'K': mes,
            'L': '',
            'M': '',
            'N': ''
        })

        ## Caso Eureca
    for i in range(len(lista_final_programas_eureca)):
        
       sheet_mapeamento.append({  
          'A': 'Eureca',
          'B': "Consta no nome do programa",
          'C': tipos_oportunidades_eureca[i],
          'D': '',
          'E': lista_final_programas_eureca[i],
          'F': semestre,
          'G': anos_programas_eureca[i],
          'H': '',
          'I': '',
          'J': lista_final_links_eureca[i],
          'K': mes,
          'L': '',
          'M': '',
          'N': ''
        })

        ## Caso Across
    for i in range(len(lista_final_programas_across)):
        
        sheet_mapeamento.append({
            'A': 'Across',
            'B': lista_final_clientes_across[i],
            'C': tipos_oportunidades_across[i],
            'D': '',
            'E': lista_final_programas_across[i],
            'F': semestre,
            'G': anos_programas_across[i],
            'H': '',
            'I': '',
            'J': links_programas_across[i],
            'K': mes,
            'L': '',
            'M': '',
            'N': ''
        })


        ## Caso MatchBox
    for i in range(len(lista_final_programas_matchbox)):
        
        sheet_mapeamento.append({
            'A': 'MatchBox',
            'B': "Consta no nome do programa",
            'C': tipos_oportunidades_matchbox[i],
            'D': '',
            'E': lista_final_programas_matchbox[i],
            'F': semestre,
            'G': anos_programas_matchbox[i],
            'H': '',
            'I': '',
            'J': lista_final_links_matchbox[i],
            'K': mes,
            'L': '',
            'M': '',
            'N': ''
        })

        ### 99Jobs
    for i in range(len(lista_final_programas_99)):
    
     sheet_mapeamento.append({
        'A': '99jobs',
        'B': lista_final_clientes_99[i],
        'C': tipos_oportunidades_99[i],
        'D': '',
        'E': lista_final_programas_99[i],
        'F': semestre,
        'G': anos_programas_99[i],
        'H': '',
        'I': '',
        'J': lista_final_links_99[i],
        'K': mes,
        'L': '',
        'M': '',
        'N': ''
    })
     
    #  SHARE RH

    for i in range(len(lista_final_programas_share)):
        sheet_mapeamento.append({
            
            'A': 'ShareRH',
            'B': lista_final_clientes_share[i],
            'C': tipos_oportunidades_share[i],
            'D': '',
            'E': lista_final_programas_share[i],
            'F': semestre,
            'G': anos_programas_share[i],
            'H': '',
            'I': '',
            'J': lista_final_links_share[i],
            'K': mes,
            'L': '',
            'M': '',
            'N': ''
    })

    if "Sheet" in workbook.sheetnames:
        workbook.remove_sheet(workbook["Sheet"])

    workbook.save(filename=f"Mapeamento de Concorrência {hoje}.xlsx")

tela_inicial()




